"""
Excel Processor Application with CIP-Signal Automation
"""

from flask import Flask, render_template, request, send_file, jsonify, session
import io
import pandas as pd
from converters import process_excel, get_process_display_name, get_process_filename
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
import threading
import time
import json
import os
import tempfile

# Import the automation module
try:
    from selenium_automation import login_and_navigate
    AUTOMATION_AVAILABLE = True
    print("✓ Selenium automation is AVAILABLE")
except ImportError as e:
    print(f"✗ Selenium automation not available: {e}")
    print("To enable automation: pip install selenium webdriver-manager")
    AUTOMATION_AVAILABLE = False

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
app.config['SECRET_KEY'] = os.urandom(24)  # For session management

# List of available processes
PROCESSES = [
    ('dbk_disbursement', 'DBK Disbursement'),
    ('dbk_pendency', 'DBK Pendency'),
    ('brc', 'BRC'),
    ('irm', 'IRM'),
    ('igst_scroll', 'IGST Scroll'),
    ('rodtep_scroll', 'RODTEP Scroll'),
    ('rodtep_scrip', 'RODTEP Scrip'),
]

# Fixed credentials for CIP-Signal
FIXED_CIP_USERNAME = "signal@master.com"
FIXED_CIP_PASSWORD = "qwerty"

def run_cip_automation_background(process_type, iec_number, file_path, brc_type=None):
    """Run CIP automation in background with fixed credentials and specific process"""
    try:
        print(f"\n{'='*60}")
        print("STARTING CIP-SIGNAL AUTOMATION")
        print(f"{'='*60}")
        print(f"Username: {FIXED_CIP_USERNAME}")
        print(f"Process: {process_type}")
        print(f"IEC Number: {iec_number if iec_number else 'Not provided'}")
        if process_type == 'brc' and brc_type:
            print(f"BRC Type: {brc_type}")
        print(f"File to upload: {file_path}")
        
        if AUTOMATION_AVAILABLE:
            # Use the new function that accepts process_type, iec_number, file path, and brc_type
            result = login_and_navigate(FIXED_CIP_USERNAME, FIXED_CIP_PASSWORD, process_type, iec_number, file_path, brc_type)
            if result['success']:
                print(f"✓ Automation SUCCESS: {result['message']}")
            else:
                print(f"✗ Automation FAILED: {result['message']}")
        else:
            print("✗ Automation not available. Install: pip install selenium webdriver-manager")
            
        print(f"{'='*60}\n")
        
        # Clean up the temporary file after automation
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Cleaned up temporary file: {file_path}")
        except Exception as e:
            print(f"Warning: Could not delete temp file: {e}")
            
    except Exception as e:
        print(f"✗ Error in background automation: {e}")
        import traceback
        traceback.print_exc()

def allowed_file(filename):
    """Check if file has allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

@app.route('/')
def index():
    """Render main page"""
    return render_template('index.html', processes=PROCESSES)

@app.route('/process', methods=['POST'])
def process_files():
    """Handle file processing"""
    temp_file_path = None
    try:
        # Get selected process
        process_type = request.form.get('process_type')
        if not process_type:
            return jsonify({'error': 'Please select a process type'}), 400
        
        # Get IEC Number (optional)
        iec_number = request.form.get('iec_number', '').strip()
        
        # Get BRC Type if BRC process is selected
        brc_type = None
        if process_type == 'brc':
            brc_type = request.form.get('brc_type')
            if not brc_type:
                return jsonify({'error': 'Please select a BRC Type'}), 400
        
        # Get files
        if 'files[]' not in request.files:
            return jsonify({'error': 'No files selected'}), 400
        
        files = request.files.getlist('files[]')
        if not files or files[0].filename == '':
            return jsonify({'error': 'No files selected'}), 400
        
        # Store file data
        file_data_list = []
        for file in files:
            if file and allowed_file(file.filename):
                filename = file.filename
                file_data = file.read()
                file_data_list.append((filename, file_data))
            else:
                return jsonify({'error': 'Invalid file type. Only .xlsx and .xls files allowed.'}), 400
        
        # Validate file count based on process type
        if process_type in ['dbk_disbursement', 'dbk_pendency', 'brc']:
            if len(file_data_list) < 1:
                return jsonify({'error': 'Please select at least one file'}), 400
        else:
            if len(file_data_list) != 1:
                return jsonify({'error': 'Please select exactly one file for this process'}), 400
        
        # Process the Excel file(s)
        try:
            processed_df = process_excel(process_type, file_data_list, brc_type)
        except ImportError as e:
            if 'xlrd' in str(e):
                return jsonify({
                    'error': 'Missing dependency: xlrd library is required to read .xls files. '
                            'Please install it using: pip install xlrd'
                }), 500
            else:
                raise e
        
        if processed_df.empty:
            return jsonify({'error': 'Processed data is empty. No data to export.'}), 400
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        
        # Write data row by row
        for row_idx, row in processed_df.iterrows():
            excel_row = row_idx + 1
            
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=excel_row, column=col_idx + 1)
                
                if pd.isna(value):
                    cell.value = None
                else:
                    cell.value = value
        
        # Apply number formatting to numeric columns based on process type
        if process_type == 'dbk_disbursement':
            # DBK Disbursement formatting
            for row in range(5, len(processed_df) + 1):
                # Column A: S No.
                cell_a = ws.cell(row=row, column=1)
                if cell_a.value is not None:
                    try:
                        cell_a.value = int(cell_a.value)
                    except:
                        pass
                
                # Column C: Shipping Bill No.
                cell_c = ws.cell(row=row, column=3)
                if cell_c.value is not None:
                    try:
                        cell_c.value = int(cell_c.value)
                    except:
                        pass
                
                # Column E: Scroll No.
                cell_e = ws.cell(row=row, column=5)
                if cell_e.value is not None:
                    try:
                        cell_e.value = int(cell_e.value)
                    except:
                        pass
                
                # Column I: Amount
                cell_i = ws.cell(row=row, column=9)
                if cell_i.value is not None:
                    try:
                        cell_i.value = int(cell_i.value)
                    except:
                        pass
            
            # Set column widths for DBK Disbursement
            column_widths = {
                'A': 8,   # S No.
                'B': 10,  # Port
                'C': 18,  # Shipping Bill No.
                'D': 18,  # Shipping Bill Date
                'E': 12,  # Scroll No.
                'F': 12,  # Scroll Date
                'G': 10,  # Drawback
                'H': 8,   # STR
                'I': 12,  # Amount
            }
            
        elif process_type == 'dbk_pendency':
            # DBK Pendency formatting
            for row in range(5, len(processed_df) + 1):
                # Column A: S No.
                cell_a = ws.cell(row=row, column=1)
                if cell_a.value is not None:
                    try:
                        cell_a.value = int(cell_a.value)
                    except:
                        pass
                
                # Column B: Shipping Bill No.
                cell_b = ws.cell(row=row, column=2)
                if cell_b.value is not None:
                    try:
                        cell_b.value = int(cell_b.value)
                    except:
                        pass
                
                # Column E: Amount
                cell_e = ws.cell(row=row, column=5)
                if cell_e.value is not None:
                    try:
                        cell_e.value = int(cell_e.value)
                    except:
                        pass
            
            # Set column widths for DBK Pendency
            column_widths = {
                'A': 8,   # S No.
                'B': 18,  # Shipping Bill No.
                'C': 18,  # Shipping Bill Date
                'D': 12,  # LEO Date
                'E': 12,  # Amount
                'F': 15,  # Current Queue
            }
        
        elif process_type == 'brc':
            # BRC formatting - 12 columns
            for row in range(4, len(processed_df) + 1):  # Data starts from row 4
                # Column D: Bill ID (convert to number)
                cell_d = ws.cell(row=row, column=4)
                if cell_d.value is not None and str(cell_d.value).strip():
                    try:
                        cell_d.value = int(float(cell_d.value))
                    except:
                        pass
                
                # Column E: SHB No (convert to number)
                cell_e = ws.cell(row=row, column=5)
                if cell_e.value is not None and str(cell_e.value).strip():
                    try:
                        cell_e.value = int(float(cell_e.value))
                    except:
                        pass
                
                # Column H: Realised Value (convert to number)
                cell_h = ws.cell(row=row, column=8)
                if cell_h.value is not None and str(cell_h.value).strip():
                    try:
                        cell_h.value = float(cell_h.value)
                        # Format as number with 2 decimal places
                        cell_h.number_format = '#,##0.00'
                    except:
                        pass
            
            # Set column widths for BRC
            column_widths = {
                'A': 20,  # BRC Number
                'B': 12,  # BRC Date
                'C': 12,  # BRC Status
                'D': 12,  # Bill ID
                'E': 12,  # SHB No
                'F': 30,  # SHB Port
                'G': 12,  # SHB Date
                'H': 15,  # Realised Value
                'I': 10,  # Currency
                'J': 18,  # Date of Realization
                'K': 20,  # BRC Utlisation Status
                'L': 10,  # BRC Lot
            }
        
        elif process_type == 'igst_scroll':
            # IGST Scroll formatting - 10 columns
            for row in range(7, len(processed_df) + 1):
                # Column A: S No.
                cell_a = ws.cell(row=row, column=1)
                if cell_a.value is not None:
                    try:
                        cell_a.value = int(cell_a.value)
                    except:
                        pass
                
                # Column B: Shipping Bill No.
                cell_b = ws.cell(row=row, column=2)
                if cell_b.value is not None:
                    try:
                        cell_b.value = int(cell_b.value)
                    except:
                        pass
                
                # Column F: Scroll Amount
                cell_f = ws.cell(row=row, column=6)
                if cell_f.value is not None:
                    try:
                        cell_f.value = float(cell_f.value)
                    except:
                        pass
            
            # Set column widths for IGST Scroll (10 columns: A-J)
            column_widths = {
                'A': 8,   # S No.
                'B': 18,  # Shipping Bill No.
                'C': 18,  # Shipping Bill Date
                'D': 18,  # IGST Scroll Number
                'E': 18,  # IGST Scroll Date
                'F': 15,  # Scroll Amount(INR)
                'G': 12,  # Scroll Status At PFMS
                'H': 12,  # Scroll Status At PAO
                'I': 15,  # Bank Response Code
                'J': 20,  # Bank Transaction Details
            }
        
        elif process_type == 'rodtep_scroll':
            # RODTEP Scroll formatting
            for row in range(4, len(processed_df) + 1):
                # Column A: Sr. No.
                cell_a = ws.cell(row=row, column=1)
                if cell_a.value is not None:
                    try:
                        cell_a.value = int(cell_a.value)
                    except:
                        pass
                
                # Column B: SHB No
                cell_b = ws.cell(row=row, column=2)
                if cell_b.value is not None:
                    try:
                        cell_b.value = int(cell_b.value)
                    except:
                        pass
                
                # Column E: Scroll Amt
                cell_e = ws.cell(row=row, column=6)
                if cell_e.value is not None:
                    try:
                        cell_e.value = float(cell_e.value)
                    except:
                        pass
            
            # Set column widths for RODTEP Scroll
            column_widths = {
                'A': 8,   # Sr. No.
                'B': 15,  # SHB No
                'C': 12,  # Date
                'D': 15,  # Scroll No
                'E': 12,  # Scroll Date
                'F': 12,  # Scroll Amt
                'G': 10,  # Port
            }
        
        elif process_type == 'rodtep_scrip':
            # RODTEP Scrip formatting - 13 columns (A-M)
            for row in range(4, len(processed_df) + 1):  # Data starts from row 4 (headers at row 3)
                # Column A: Sr. No.
                cell_a = ws.cell(row=row, column=1)
                if cell_a.value is not None and str(cell_a.value).strip():
                    try:
                        cell_a.value = int(float(cell_a.value))
                    except:
                        pass
                
                # Column B: SCROLL NUMBER
                cell_b = ws.cell(row=row, column=2)
                if cell_b.value is not None and str(cell_b.value).strip():
                    try:
                        cell_b.value = int(float(cell_b.value))
                    except:
                        pass
                
                # Column C: SB NUMBER
                cell_c = ws.cell(row=row, column=3)
                if cell_c.value is not None and str(cell_c.value).strip():
                    try:
                        cell_c.value = int(float(cell_c.value))
                    except:
                        pass
                
                # Column F: SCRIP NUMBER
                cell_f = ws.cell(row=row, column=6)
                if cell_f.value is not None and str(cell_f.value).strip():
                    try:
                        cell_f.value = int(float(cell_f.value))
                    except:
                        pass
                
                # Column I: SCRIP ISSUE AMOUNT
                cell_i = ws.cell(row=row, column=9)
                if cell_i.value is not None and str(cell_i.value).strip():
                    try:
                        cell_i.value = float(cell_i.value)
                        cell_i.number_format = '#,##0.00'
                    except:
                        pass
                
                # Column J: SCRIP BALANCE AMOUNT
                cell_j = ws.cell(row=row, column=10)
                if cell_j.value is not None and str(cell_j.value).strip():
                    try:
                        cell_j.value = float(cell_j.value)
                        cell_j.number_format = '#,##0.00'
                    except:
                        pass
            
            # Set column widths for RODTEP Scrip (13 columns: A-M)
            column_widths = {
                'A': 8,   # Sr. No.
                'B': 15,  # SCROLL NUMBER
                'C': 15,  # SB NUMBER
                'D': 12,  # SB DATE
                'E': 12,  # SB AMOUNT
                'F': 15,  # SCRIP NUMBER
                'G': 15,  # SCRIP ISSUE DATE
                'H': 15,  # SCRIP EXPIRY DATE
                'I': 18,  # SCRIP ISSUE AMOUNT
                'J': 18,  # SCRIP BALANCE AMOUNT
                'K': 18,  # SCRIP TRANSFER DATE
                'L': 15,  # SCRIP STATUS
                'M': 20,  # Application Ref. No
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Center align all data cells for RODTEP Scrip (columns A-M, starting from row 4)
            # Except for amount columns (I and J) which are right-aligned
            for row in range(4, len(processed_df) + 1):  # Data starts from row 4
                for col in range(1, 14):  # Columns A-M (1-13)
                    cell = ws.cell(row=row, column=col)
                    # Don't center align amounts (columns I and J) - right align for numbers
                    if col in [9, 10]:  # Columns I and J (SCRIP ISSUE AMOUNT, SCRIP BALANCE AMOUNT)
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='center')   
        else:
            # Default column widths for other processes
            column_widths = {
                'A': 15,
                'B': 15,
                'C': 15,
                'D': 15,
                'E': 15,
                'F': 15,
                'G': 15,
                'H': 15,
                'I': 15,
                'J': 15,  # Added for processes that might have 10 columns
                'K': 15,  # Added for BRC (11 columns)
                'L': 15,  # Added for BRC (12 columns)
                'M': 15,  # Added for RODTEP Scrip (13 columns)
            }
        
        # Apply column widths (for processes other than rodtep_scrip which already applied them)
        if process_type != 'rodtep_scrip':
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
        
        # Center align all data cells for other processes
        center_alignment = Alignment(horizontal='center')
        
        if process_type == 'dbk_pendency':
            # For DBK Pendency, align data from row 5 onwards (columns A-F)
            for row in range(5, len(processed_df) + 1):
                for col in range(1, 7):  # Columns A-F
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
        elif process_type == 'brc':
            # For BRC, align data from row 4 onwards (columns A-L)
            for row in range(4, len(processed_df) + 1):
                for col in range(1, 13):  # Columns A-L (1-12)
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
        elif process_type == 'igst_scroll':
            # For IGST Scroll, align data from row 7 onwards (columns A-J)
            for row in range(7, len(processed_df) + 1):
                for col in range(1, 11):  # Columns 1 to 10 (A-J)
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
        elif process_type == 'rodtep_scroll':
            # For RODTEP Scroll, align data from row 4 onwards (columns A-G)
            for row in range(4, len(processed_df) + 1):
                for col in range(1, 8):  # Columns A-G
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
        elif process_type == 'dbk_disbursement':
            # For DBK Disbursement, align data from row 5 onwards (columns A-I)
            for row in range(5, len(processed_df) + 1):
                for col in range(1, 10):  # Columns A-I
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
        # Note: rodtep_scrip alignment is already handled above
        
        # Save workbook to bytes first
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        process_filename = get_process_filename(process_type)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if process_type in ['dbk_disbursement', 'dbk_pendency', 'brc']:
            download_name = f"merged_{process_filename}_{timestamp}.xlsx"
        else:
            download_name = f"{process_filename}_{timestamp}.xlsx"
        
        # Create a temporary file to save the processed Excel
        temp_dir = tempfile.mkdtemp()
        temp_file_path = os.path.join(temp_dir, download_name)
        
        # Save the workbook to the temp file
        wb.save(temp_file_path)
        
        print(f"\n✓ File saved to temp location: {temp_file_path}")
        
        # Start automation in background with the temp file path
        print(f"Starting CIP-Signal automation for process: {process_type}...")
        
        # Run automation in a separate thread
        automation_thread = threading.Thread(
            target=run_cip_automation_background, 
            args=(process_type, iec_number, temp_file_path, brc_type)  # Pass brc_type here
        )
        automation_thread.daemon = True
        automation_thread.start()
        
        # Send the file for download
        return send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except ImportError as e:
        if 'xlrd' in str(e):
            return jsonify({
                'error': 'Missing dependency: xlrd library is required to read .xls files. '
                        'Please install it using: pip install xlrd'
            }), 500
        else:
            # Clean up temp file if it was created
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass
            return jsonify({'error': str(e)}), 500
    except Exception as e:
        # Clean up temp file if it was created
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except:
                pass
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("="*60)
    print("EXCEL PROCESSOR WITH CIP-SIGNAL AUTOMATION")
    print("="*60)
    print(f"Fixed CIP-Signal credentials:")
    print(f"  Username: {FIXED_CIP_USERNAME}")
    print(f"  Password: {FIXED_CIP_PASSWORD}")
    print("="*60)
    
    if AUTOMATION_AVAILABLE:
        print("✓ Selenium automation: ENABLED")
        # print("  Automation will run automatically after each file download")
        # print("  - Browser will open VISIBLY (not in background)")
        # print("  - Will upload the processed file to CIP-Signal portal")
        # print("  - Will select IEC if provided")
        # print("  - For BRC: Will select BRC type (FOC/INV) before IEC selection")
    else:
        print("✗ Selenium automation: NOT AVAILABLE")
        # print("  To enable automation, install:")
        # print("    pip install selenium webdriver-manager")
    
    print("\nStarting Flask server...")
    app.run(host='0.0.0.0', port=5000, debug=True)